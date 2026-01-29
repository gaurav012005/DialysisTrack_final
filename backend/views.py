from rest_framework import permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Count, Sum, Avg, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json
from io import BytesIO

# Optional imports for PDF and Excel functionality
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def reports_list(request):
    """List available report endpoints"""
    base_url = request.build_absolute_uri('/api/reports/')
    
    reports = {
        'message': 'Available Reports',
        'endpoints': {
            'dashboard_stats': f'{base_url}dashboard-stats/',
            'patient_reports': f'{base_url}patients/',
            'queue_reports': f'{base_url}queue/',
            'machine_reports': f'{base_url}machines/',
            'staff_reports': f'{base_url}staff/',
            'export_reports': f'{base_url}export/'
        }
    }
    
    return Response(reports)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def dashboard_stats(request):
    """Get overall dashboard statistics"""
    from patients.models import Patient
    from dialysis_queue.models import Queue
    from machines.models import DialysisMachine
    from appointments.models import Appointment
    
    today = datetime.now().date()
    
    # Patient Statistics
    total_patients = Patient.objects.filter(is_active=True).count()
    emergency_patients = Patient.objects.filter(is_emergency=True, is_active=True).count()
    
    # Queue Statistics
    queue_stats = Queue.objects.filter(
        check_in_time__date=today
    ).aggregate(
        total_waiting=Count('id', filter=Q(status='waiting')),
        total_in_progress=Count('id', filter=Q(status='in_progress')),
        total_completed=Count('id', filter=Q(status='completed'))
    )
    
    # Machine Statistics
    machine_stats = DialysisMachine.objects.filter(is_active=True).aggregate(
        total_machines=Count('id'),
        available_machines=Count('id', filter=Q(status='available')),
        in_use_machines=Count('id', filter=Q(status='in_use'))
    )
    
    # Appointment Statistics
    appointment_stats = Appointment.objects.filter(
        appointment_date=today
    ).aggregate(
        total_appointments=Count('id'),
        completed_appointments=Count('id', filter=Q(status='completed'))
    )
    
    stats = {
        'patients': {
            'total': total_patients,
            'emergency': emergency_patients,
            'active_today': Queue.objects.filter(check_in_time__date=today).count()
        },
        'queue': queue_stats,
        'machines': machine_stats,
        'appointments': appointment_stats,
        'date': today
    }
    
    return Response(stats)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def patient_reports(request):
    """Generate patient reports"""
    from patients.models import Patient
    from patients.serializers import PatientSerializer
    
    report_type = request.GET.get('type', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    patients = Patient.objects.all()
    
    if report_type == 'emergency':
        patients = patients.filter(is_emergency=True)
    elif report_type == 'active':
        patients = patients.filter(is_active=True)
    
    if start_date and end_date:
        patients = patients.filter(created_at__date__range=[start_date, end_date])
    
    serializer = PatientSerializer(patients, many=True)
    
    return Response({
        'report_type': report_type,
        'total_patients': patients.count(),
        'patients': serializer.data
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def queue_reports(request):
    """Generate queue performance reports"""
    from dialysis_queue.models import Queue
    
    report_type = request.GET.get('type', 'daily')
    date_param = request.GET.get('date', datetime.now().date())
    
    if report_type == 'daily':
        queues = Queue.objects.filter(check_in_time__date=date_param)
    elif report_type == 'weekly':
        start_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        end_date = start_date + timedelta(days=6)
        queues = Queue.objects.filter(check_in_time__date__range=[start_date, end_date])
    elif report_type == 'monthly':
        year, month = map(int, date_param.split('-'))
        queues = Queue.objects.filter(
            check_in_time__year=year,
            check_in_time__month=month
        )
    
    stats = queues.aggregate(
        total_patients=Count('id'),
        avg_wait_time=Avg('estimated_wait_time'),
        completed_sessions=Count('id', filter=Q(status='completed')),
        emergency_cases=Count('id', filter=Q(priority='emergency'))
    )
    
    return Response({
        'report_type': report_type,
        'period': date_param,
        'statistics': stats,
        'total_records': queues.count()
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def machine_utilization(request):
    """Generate machine utilization reports"""
    from machines.models import DialysisMachine, MaintenanceLog
    
    report_type = request.GET.get('type', 'overview')
    
    if report_type == 'overview':
        machines = DialysisMachine.objects.filter(is_active=True)
        
        utilization_data = []
        for machine in machines:
            utilization_rate = (machine.total_sessions / 100) * 100 if machine.total_sessions > 0 else 0
            
            utilization_data.append({
                'machine_id': machine.machine_id,
                'machine_name': machine.name,
                'status': machine.status,
                'total_sessions': machine.total_sessions,
                'total_hours': float(machine.total_operating_hours),
                'utilization_rate': round(utilization_rate, 2),
                'needs_maintenance': machine.needs_maintenance
            })
        
        return Response({
            'report_type': 'machine_utilization',
            'machines': utilization_data
        })
    
    elif report_type == 'maintenance':
        maintenance_logs = MaintenanceLog.objects.all().order_by('-maintenance_date')[:50]
        
        from machines.serializers import MaintenanceLogSerializer
        serializer = MaintenanceLogSerializer(maintenance_logs, many=True)
        
        return Response({
            'report_type': 'maintenance_history',
            'maintenance_logs': serializer.data
        })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def staff_performance(request):
    """Generate staff performance reports"""
    from staff.models import StaffAttendance
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    report_type = request.GET.get('type', 'attendance')
    month = request.GET.get('month', datetime.now().month)
    year = request.GET.get('year', datetime.now().year)
    
    if report_type == 'attendance':
        attendance_data = StaffAttendance.objects.filter(
            date__year=year,
            date__month=month
        ).values('staff').annotate(
            total_days=Count('id'),
            present_days=Count('id', filter=Q(status='present')),
            absent_days=Count('id', filter=Q(status='absent')),
            late_days=Count('id', filter=Q(status='late'))
        )
        
        performance_data = []
        for data in attendance_data:
            staff = User.objects.get(id=data['staff'])
            attendance_rate = (data['present_days'] / data['total_days'] * 100) if data['total_days'] > 0 else 0
            
            performance_data.append({
                'staff_name': staff.get_full_name(),
                'role': staff.role,
                'total_days': data['total_days'],
                'present_days': data['present_days'],
                'absent_days': data['absent_days'],
                'late_days': data['late_days'],
                'attendance_rate': round(attendance_rate, 2)
            })
        
        return Response({
            'report_type': 'staff_attendance',
            'month': month,
            'year': year,
            'performance_data': performance_data
        })

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def export_report(request):
    """Export report data in multiple formats"""
    report_type = request.GET.get('type', 'patients')
    format_type = request.GET.get('format', 'csv')
    
    # Check format availability
    if format_type == 'pdf' and not PDF_AVAILABLE:
        return Response({'error': 'PDF export not available. Please install reportlab.'}, status=400)
    if format_type == 'excel' and not EXCEL_AVAILABLE:
        return Response({'error': 'Excel export not available. Please install openpyxl.'}, status=400)
    
    if report_type == 'patients':
        return export_patients_report(format_type)
    elif report_type == 'queue':
        return export_queue_report(format_type)
    elif report_type == 'machines':
        return export_machines_report(format_type)
    
    return Response({'error': 'Invalid report type'}, status=status.HTTP_400_BAD_REQUEST)

def export_patients_report(format_type):
    """Export patients data in specified format"""
    from .simple_exports import export_patients_csv
    
    if format_type == 'csv':
        return export_patients_csv()
    
    from patients.models import Patient
    patients = Patient.objects.all().order_by('patient_id')
    
    if format_type == 'excel':
        if not EXCEL_AVAILABLE:
            return HttpResponse('Excel export not available. Please install openpyxl.', status=400)
        
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Patients Report'
        
        # Headers
        headers = ['Patient ID', 'Name', 'Gender', 'DOB', 'Phone', 'Email', 'Blood Type', 'Emergency', 'Status']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, patient in enumerate(patients, 2):
            worksheet.cell(row=row, column=1, value=patient.patient_id)
            worksheet.cell(row=row, column=2, value=f"{patient.first_name} {patient.last_name}")
            worksheet.cell(row=row, column=3, value=patient.gender)
            worksheet.cell(row=row, column=4, value=patient.date_of_birth.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=5, value=patient.phone_number)
            worksheet.cell(row=row, column=6, value=patient.email or 'N/A')
            worksheet.cell(row=row, column=7, value=patient.blood_type or 'N/A')
            worksheet.cell(row=row, column=8, value='Yes' if patient.is_emergency else 'No')
            worksheet.cell(row=row, column=9, value='Active' if patient.is_active else 'Inactive')
        
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="patients_report.xlsx"'
        return response
    
    elif format_type == 'pdf':
        if not PDF_AVAILABLE:
            return HttpResponse('PDF export not available. Please install reportlab.', status=400)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        title = Paragraph('Patients Report', title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        date_para = Paragraph(f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}', date_style)
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Patient ID', 'Name', 'Gender', 'Phone', 'Emergency', 'Status']]
        
        for patient in patients:
            data.append([
                patient.patient_id,
                f"{patient.first_name} {patient.last_name}",
                patient.gender,
                patient.phone_number,
                'Yes' if patient.is_emergency else 'No',
                'Active' if patient.is_active else 'Inactive'
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="patients_report.pdf"'
        return response

def export_queue_report(format_type):
    """Export queue data"""
    from dialysis_queue.models import Queue
    from .simple_exports import export_queue_csv
    
    if format_type == 'csv':
        return export_queue_csv()
    
    elif format_type == 'pdf':
        if not PDF_AVAILABLE:
            return HttpResponse('PDF export not available. Please install reportlab.', status=400)
        
        queues = Queue.objects.all().order_by('-check_in_time')
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        title = Paragraph('Queue Report', title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        date_para = Paragraph(f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}', date_style)
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Patient', 'Machine', 'Status', 'Priority', 'Check-in Time']]
        
        for queue in queues[:50]:  # Limit to 50 records
            data.append([
                f"{queue.patient.first_name} {queue.patient.last_name}",
                queue.machine.name if queue.machine else 'N/A',
                queue.status,
                queue.priority,
                queue.check_in_time.strftime('%Y-%m-%d %H:%M')
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="queue_report.pdf"'
        return response
    
    return HttpResponse('Format not supported', status=400)

def export_machines_report(format_type):
    """Export machines data"""
    from machines.models import DialysisMachine
    from .simple_exports import export_machines_csv
    
    if format_type == 'csv':
        return export_machines_csv()
    
    elif format_type == 'pdf':
        if not PDF_AVAILABLE:
            return HttpResponse('PDF export not available. Please install reportlab.', status=400)
        
        machines = DialysisMachine.objects.all().order_by('machine_id')
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        title = Paragraph('Machines Report', title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        date_para = Paragraph(f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}', date_style)
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Machine ID', 'Name', 'Status', 'Sessions', 'Hours', 'Maintenance']]
        
        for machine in machines:
            data.append([
                machine.machine_id,
                machine.name,
                machine.status,
                str(machine.total_sessions),
                f"{machine.total_operating_hours:.1f}",
                'Yes' if machine.needs_maintenance else 'No'
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="machines_report.pdf"'
        return response
    
    return HttpResponse('Format not supported', status=400)